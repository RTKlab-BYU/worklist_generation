from openpyxl.styles import Border, Side, Font, Protection

from metadata_capture.excel_utils.excel_utils import copy_template_with_datetime
from metadata_capture.project_dataclasses.project_outline import ProjectOutline
from abc import ABC, abstractmethod
from openpyxl import load_workbook
from copy import copy


class ExcelFileGenerator(ABC):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        self.source_file = source_file
        self.project_outline = project_outline
        self.excel_path: str = self._make_copy()
        self.wb = load_workbook(self.excel_path, keep_vba=True)
        self.ws = self.wb["DataEntry"]

    def _make_copy(self) -> str:
        return copy_template_with_datetime(filename=self.source_file, new_name=self.project_outline.name)

    def generate_file(self) -> str:
        self.ws.cell(row=1, column=1).value = self.project_outline.name
        self._make_changes()
        self.ws.protection.sheet = True
        self.wb.save(self.excel_path)
        return self.excel_path

    @abstractmethod
    def _make_changes(self) -> None:
        pass



class BasicFileGenerator(ExcelFileGenerator):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        super().__init__(source_file, project_outline)
        raise NotImplementedError

    def _make_changes(self) -> None:
        pass



class AdvancedFileGenerator(ExcelFileGenerator):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        super().__init__(source_file, project_outline)

    def _make_changes(self) -> None:
        """
        Efficiently copies formatting from the format pattern column to new columns
        and inserts group names from the project outline.
        """
        # Step 1: Load format pattern once (bug fixed: increment row!)
        format_pattern = self._get_format_pattern()

        # Step 2: Pre-copy all style attributes ONCE to avoid repetitive .cell() + copy() calls
        style_attrs = [
            {
                "font": copy(cell.font),
                "border": Border(
                                left=Side(style="thin", color="000000"),
                                right=Side(style="thin", color="000000"),
                                top=Side(style="thin", color="000000"),
                                bottom=Side(style="thin", color="000000"),
                                ),
                "fill": copy(cell.fill),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
                "alignment": copy(cell.alignment),
            }
            for cell in format_pattern
        ]

        # Step 3: Write each group's name and apply formatting efficiently
        cur_col = 3
        for group in self.project_outline.groups:
            # Set group name in header row
            self.ws.cell(row=8, column=cur_col).value = group
            self.ws.cell(row=8, column=cur_col).font = Font(size=20, bold=True)
            # Apply pre-copied styles down the column
            for offset, style in enumerate(style_attrs):
                row_idx = 9 + offset
                target = self.ws.cell(row=row_idx, column=cur_col)
                # Apply style attributes
                target.font = style["font"]
                target.border = style["border"]
                target.fill = style["fill"]
                target.number_format = style["number_format"]
                target.protection = style["protection"]
                target.alignment = style["alignment"]
                target.protection = Protection(locked=False)


            cur_col += 1

    def _get_format_pattern(self):
        cell_format_pattern = []
        row = 9
        COL = 3
        while True:
            cell = self.ws.cell(row=row, column=COL)
            val = cell.value
            if val == "END_OF_FILE":
                break
            if row > 200:
                raise ValueError("INFINITE LOOP. No 'END_OF_FILE' value found in column 3")

            cell_format_pattern.append(cell)
            row += 1
        return cell_format_pattern


class CustomFileGenerator(ExcelFileGenerator):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        super().__init__(source_file, project_outline)
        raise NotImplementedError

    def _make_changes(self) -> None:
        pass

