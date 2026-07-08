#!/usr/bin/env python3

import sys
import os
import platform
import subprocess
import argparse
from pathlib import Path

# Stage 1 imports
from metadata_capture.cli_utils.cli_prompts import get_project_outline
from metadata_capture.project_dataclasses.project_type import ProjectType
from metadata_capture.excel_utils.excel_file_generator import AdvancedFileGenerator
from metadata_capture.excel_utils.excel_utils import open_excel_sheet
from metadata_capture.vars import TEMPLATES

# Stage 2 imports
from metadata_capture.excel_utils.excel_file_parser import AdvancedFileParser
from metadata_capture.database_utils.upload_to_sqlite import upload_to_sqlite
from metadata_capture.excel_utils.fill_conditions import fill_conditions_in_worklist

# Stage 3 imports
import openpyxl
import worklist_classes.main as worklist_main
from metadata_capture.sdrf_generator import generate_sdrf_for_worklist


def parse_args():
    parser = argparse.ArgumentParser(
        prog="run.py",
        usage="""
python run.py -s 1
python run.py -s 2 -m <metadata_excel_path>
python run.py -s 3 -w <worklist_excel_path> -o <output_dir>
""",
        description="LC/MS Worklist Pipeline",
        formatter_class=argparse.RawTextHelpFormatter
    )

    parser.add_argument(
        "-s", "--stage",
        required=True,
        choices=["1", "2", "3"],
        help="Pipeline stage"
    )

    parser.add_argument(
        "-m", "--metadata",
        type=Path,
        help="Metadata Excel file (stage 2)"
    )

    parser.add_argument(
        "-w", "--worklist",
        type=Path,
        help="Worklist Excel file (stage 3)"
    )

    parser.add_argument(
        "-o", "--output",
        type=Path,
        help="Output directory (stage 3)"
    )

    # parser.add_argument(
    #     "-p", "--project_id",
    #     type=int,
    #     help="Project ID (stage 4)"
    # )

    return parser.parse_args()


def validate_args(args):
    if args.stage == "2" and not args.metadata:
        raise SystemExit("Stage 2 requires -m / --metadata")

    if args.stage == "3":
        if not args.worklist or not args.output:
            raise SystemExit("Stage 3 requires -w and -o")

    # if args.stage == "4" and not args.project_id:
    #     raise SystemExit("Stage 4 requires -p / --project_id")


# Helper Functions
def auto_open_file(path: Path):
    """Open file cross-platform."""
    try:
        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":
            subprocess.call(["open", path])
        else:
            subprocess.call(["xdg-open", path])
    except Exception:
        print(f"Could not auto-open file: {path}")


def usage():
    print("""
USAGE:
  python run.py -s 1
      → Generate metadata Excel

  python run.py -s 2 -m <metadata_excel_path>
      → Generate filled worklist Excel and SDRF file

  python run.py -s 3 -w <worklist_excel_path> -o <output_dir>
      → Generate LC + MS worklists
""")
    sys.exit(1)


# Stage 1
def stage_1_generate_metadata():
    """
    Collect project info and generate metadata Excel.
    """
    print("\n=== STAGE 1: Generate Metadata Excel ===\n")

    project_outline = get_project_outline()
    project_type = ProjectType.ADVANCED

    source_file = TEMPLATES / "metadataTemplate8.xlsm"
    generator = AdvancedFileGenerator(
        source_file=source_file,
        project_outline=project_outline
    )
    temp_path = Path(generator.generate_file())

    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(exist_ok=True)

    output_path = temp_path.rename(output_dir / temp_path.name)

    print(f"\nMetadata Excel generated:\n  {output_path}")
    auto_open_file(output_path)

    print("\nNext step (you may copy and paste the command below):")
    print(f"  python run.py -s 2 -m {output_path}\n")

    sys.exit(0)


# Stage 2
def stage_2_generate_worklist(metadata_excel_path: Path):
    """
    Parse metadata Excel, upload to DB, generate filled worklist Excel, and generate SDRF file.
    """
    print("\n=== STAGE 2: Generate Filled Worklist ===\n")

    if not metadata_excel_path.exists():
        print(f"File not found: {metadata_excel_path}")
        sys.exit(1)

    parser = AdvancedFileParser(
        source_file=str(metadata_excel_path),
        project_outline=None  # parser reconstructs needed info internally
    )

    project_metadata = parser.parse_file()

    
    project_id = upload_to_sqlite(project_metadata)

    print(f"Project uploaded (ID: {project_id})")

    output_path = fill_conditions_in_worklist(
        project_id=project_id,
        db_path="project.db",
        template_path=str(TEMPLATES / "worklist_template_blank.xlsx"),
        output_dir="output",
        output_filename=f"worklist_conditions_project_{project_id}.xlsx"
    )

    temp_path = Path(output_path)

    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(exist_ok=True)

    output_path = temp_path.rename(output_dir / temp_path.name)

    print(f"\nMetadata captured!\nWorklist template created:\n  {output_path}")
    auto_open_file(output_path)

    print("\nNext step (you may copy and paste the command below, but remember to add your own output directory):")
    print(f"  python run.py -s 3 -w {output_path} -o <output_dir>\n")

    sys.exit(0)


def get_project_id_from_worklist(worklist_excel_path: Path) -> int:
    """
    Read the project_id back out of the 'User' sheet, cell B4, where stage 2
    writes it (fill_conditions_in_worklist). Reads via openpyxl directly
    rather than pandas, so it's unaffected by header-row offsets and by the
    sheet being protected/locked.
    """
    wb = openpyxl.load_workbook(worklist_excel_path, data_only=True)
    if "User" not in wb.sheetnames:
        raise ValueError(f"'User' sheet not found in {worklist_excel_path}")
    ws = wb["User"]
    label = ws.cell(row=4, column=1).value
    project_id = ws.cell(row=4, column=2).value
    if project_id is None or (label and "Project ID" not in str(label)):
        return None
    return int(project_id)


# Stage 3
def stage_3_generate_lcms(worklist_excel_path: Path, output_dir: Path):
    """
    Generate LC and MS worklist files, and the SDRF for the project.
    """
    print("\n=== STAGE 3: Generate LC/MS Worklists ===\n")

    if not worklist_excel_path.exists():
        print(f"File not found: {worklist_excel_path}")
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)

    project_id = get_project_id_from_worklist(worklist_excel_path)

    ms_pd, lc_pd, ms_filename, lc_filename, filenames, condition_names, rep_numbers = worklist_main.main(
        str(worklist_excel_path)
    )

    ms_path = output_dir / ms_filename
    lc_path = output_dir / lc_filename

    ms_pd.to_csv(ms_path, index=False, header=False, encoding="utf-8-sig")
    lc_pd.to_csv(lc_path, index=False, header=False, encoding="utf-8-sig")

    print("\nWorklists generated:")
    print(f"  MS → {ms_path}")
    print(f"  LC → {lc_path}")

    # --- SDRF export (moved from stage 2) ---
    if project_id is None:
        print("  SDRF → skipped (no Project ID found in worklist)")
    else:
        sdrf_path = generate_sdrf_for_worklist(
            project_id=project_id,
            filenames=filenames,
            condition_names=condition_names,
            rep_numbers=rep_numbers,
            db_path="project.db",
            output_dir=str(output_dir),
        )
        print(f"  SDRF → {sdrf_path}")

    sys.exit(0)

# # Stage 4
# def stage_4_print_sdrf(project_id: int):
#     """
#     Print the SDRF for a previously uploaded project to stdout.
#     """
#     print(f"\n=== STAGE 4: Print SDRF (Project ID: {project_id}) ===\n")
 
#     from metadata_capture.sdrf_generator import generate_sdrf_rows
 
#     headers, rows = generate_sdrf_rows(project_id=project_id, db_path="project.db")
 
#     print("\t".join(headers))
#     for row in rows:
#         print("\t".join(row))
 
#     sys.exit(0)


# Main
def main():
    args = parse_args()
    validate_args(args)

    if args.stage == "1":
        stage_1_generate_metadata()

    elif args.stage == "2":
        if not args.metadata:
            print("Stage 2 requires --metadata")
            return
        stage_2_generate_worklist(args.metadata)

    elif args.stage == "3":
        if not args.worklist or not args.output:
            print("Stage 3 requires --worklist and --output")
            return
        stage_3_generate_lcms(args.worklist, args.output)

    # elif args.stage == "4":
    #     if not args.project_id:
    #         print("Stage 4 requires --project_id")
    #         return
    #     stage_4_print_sdrf(args.project_id)
    else:
        usage()


if __name__ == "__main__":
    main()
