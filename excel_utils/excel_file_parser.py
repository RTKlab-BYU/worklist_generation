from abc import ABC, abstractmethod
from collections import defaultdict

import openpyxl

from project_dataclasses.project_metadata import ProjectMetadata
from project_dataclasses.project_outline import ProjectOutline


class ExcelFileParser(ABC):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        self.source_file = source_file
        self.project_outline = project_outline
        self.wb = openpyxl.load_workbook(self.source_file, read_only=True)
        self.ws = self.wb["DataEntry"]


    def parse_file(self) -> ProjectMetadata:
        return self._parse_file()

    @abstractmethod
    def _parse_file(self) -> ProjectMetadata:
        pass


class BasicFileParser(ExcelFileParser):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        super().__init__(source_file, project_outline)
        raise NotImplementedError

    def _parse_file(self) -> ProjectMetadata:
        pass



class AdvancedFileParser(ExcelFileParser):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        super().__init__(source_file, project_outline)

    def _parse_file(self) -> ProjectMetadata:


        self._validate_title_and_group_names()


        return self._get_metadata()


    def _get_metadata(self) -> ProjectMetadata:

        # Categories: {category: (start_row, end_row)}
        categories: dict[str, tuple[int, int]] = self._get_categories()

        # Metadata: {group: {category: {label: value}}}
        metadata: dict[str, dict[str, dict[str, str]]] = defaultdict(lambda: defaultdict(dict))

        # Independent Variables: {category: {label: [levels/values]}}
        independent_variables: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))

        for category in categories:
            start_row, stop_row = categories[category]
            for row_number in range(start_row, stop_row + 1):
                self._read_row(row_number, category, metadata, independent_variables)

        return ProjectMetadata(project_outline=self.project_outline,
                               independent_variables=independent_variables,
                               metadata=metadata)

    def _get_categories(self) -> dict[str, tuple[int, int]]:

        categories = {}

        cur_row = 9
        while True:
            cur_cat, end_row = self._get_one_cat(cur_row, categories)

            cur_row = end_row

            if not cur_cat:
                break

            if cur_row > 500:
                raise IndexError("Went > 500 on while loop in _get_categories")

        return categories


    def _get_one_cat(self, cur_row, categories) -> tuple[str, int]:
        category = self.ws.cell(cur_row, 1).value
        if category == "END_OF_FILE":
            return '', cur_row

        og_row = cur_row
        val = ''
        while val != "XXXXXXXXXX":
            val = self.ws.cell(cur_row, 1).value
            cur_row += 1
            if cur_row > 500:
                raise IndexError("Went > 500 on while loop in _get_one_cat")

        categories[category] = (og_row + 1, cur_row - 1)

        return category, cur_row


    def _read_row(self, row_number: int, cur_cat: str, metadata: dict, ind_vars: dict) -> None:

        label: str = str(self.ws.cell(row=row_number, column=2).value).strip()

        contains_ind_var = False

        values: set[str] = set()

        first_val: str = ""

        for index, group in enumerate(self.project_outline.groups):

            if cur_cat in metadata[group] and label in metadata[group][cur_cat]:
                raise ValueError("Can't have two labels that equal the same thing.")

            col = index + 3
            val = self.ws.cell(row=row_number, column=col).value

            if index == 0:
                if not val:
                    break
                first_val = val
                values.add(val)

            # THE VALUE WILL DEFAULT TO THE FIRST COLUMN IF THE OTHER COLUMNS ARE LEFT BLANK
            elif not val:
                val = first_val

            if val not in values:
                contains_ind_var = True
                values.add(val)


            # ADD IT TO METADATA
            metadata[group][cur_cat][label] = val


        if contains_ind_var:
            ind_vars[cur_cat][label] = list(values)

    def _validate_title_and_group_names(self):
        if self.ws.cell(1,1).value != self.project_outline.name:
            raise ValueError("Incorrect File Submitted. Title != Project Name")

        for index, group in enumerate(self.project_outline.groups):
            col = index + 3
            val = self.ws.cell(8, col).value
            if val != group:
                raise ValueError(f"Group Mismatch at (8, {col}); GOT: {val} != EXPECTED: {group}")


class CustomFileParser(ExcelFileParser):

    def __init__(self, source_file: str, project_outline: ProjectOutline):
        super().__init__(source_file, project_outline)
        raise NotImplementedError

    def _parse_file(self) -> ProjectMetadata:
        pass

